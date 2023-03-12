import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname": "Amit", "lastname":"Pandey", "gender":"Male"},{"firstname":"Any","lastname":"pandey","gender":"Female"}]
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\DELL\\Desktop\\PythonDemo.xlsx")  # loading workbook from the location

        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # here we want to print only the value of the row
                #  which has the name as Testcase2.
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]