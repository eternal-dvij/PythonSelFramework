import openpyxl # importing openpyxl

book = openpyxl.load_workbook("C:\\Users\\DELL\\Desktop\\PythonDemo.xlsx") # loading workbook from the location

sheet = book.active # initializing the sheet variable with the current active sheet

cell = sheet.cell(row=1,column=2) # putting the value of cell which is on row=1 and column=2
# print(cell.value) # printing the value of that cell

sheet.cell(row=2, column=2).value = "Rahul"

dict = {}
# print(sheet.cell(row=2,column=2).value)
#
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet['A5'].value)


for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        # here we want to print only the value of the row
        #  which has the name as Testcase2.
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print((dict))